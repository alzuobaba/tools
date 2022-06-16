# coding=utf-8
# openpyxl require python3

import requests
from openpyxl import load_workbook
from lxml import etree
import re


def get_cves():
    # 打开一个workbook
    wb = load_workbook(filename="cve.xlsx")

    # 获取当前活跃的worksheet,默认就是第一个worksheet
    # ws = wb.active

    # 当然也可以使用下面的方法

    # 获取所有表格(worksheet)的名字
    sheets = wb.sheetnames
    # 第一个表格的名称
    sheet_first = sheets[0]
    # 获取特定的worksheet
    ws = wb[sheet_first]

    # 获取表格所有行和列，两者都是可迭代的
    rows = list(ws.rows)
    columns = list(ws.columns)
    row = rows[1]
    column = columns[1]
    lines = [col.value for col in column[1:-1]]
    cves = dict()
    nocves = []
    for line in lines:
        # print(line)
        searchObj = re.search('cve.*\d', line, re.I | re.M)
        if searchObj:
            cves.update({line: searchObj.group()})
        else:
            nocves.append(line)
    return cves, nocves


def get_cve_errata_urls(cve_num, affect_product_cpe='cpe:/o:redhat:enterprise_linux:7'):
    ret = []
    url = "https://access.redhat.com/api/redhat_node/{}.json".format(cve_num)
    print(url)
    response = requests.request("GET", url)
    if response.status_code != 200:
        ret.append((None, None))
    else:
        res = response.json()
        objs = res['field_cve_releases_txt']['und'][0]['object']
        objs = [obj for obj in objs if obj['cpe'] == affect_product_cpe]
        if not objs:
            ret.append(("未找到RedHat7平台", None))
        for obj in objs:
            errate_state = str(obj['package']) + ':' + str(obj['state'])
            errate_url = obj['advisory'].get('url')
            ret.append((errate_state, errate_url))
    return ret


def get_errata_rpm_packages(url, platform_name="Red Hat Enterprise Linux Server 7"):
    rpm_infos_str = ''
    response = requests.request("GET", url)
    html = etree.HTML(response.text)
    title_node = html.xpath('//*[@id="packages"]/h2[text()="{}"]'.format(platform_name))
    if title_node and title_node[0].getnext().tag == 'table':
        table = title_node[0].getnext()
        rpm_infos = table.xpath('.//tr/td[@class="name"]//text()')
        rpm_infos = [str(s).strip() for s in rpm_infos]
        rpm_infos_str += '\n'.join(rpm_infos)
    rpm_infos_str += '\n'
    return rpm_infos_str


def get_errata_details(cve_num):
    ret_str = ""
    errata_urls = get_cve_errata_urls(cve_num)
    for state, url in errata_urls:
        if not state:
            ret_str += '红帽未收录\n'
        elif not url:
            ret_str += '红帽未处理: %s\n' % state
        else:
            ret_str += '包名：%s\n' % state
            ret_str += url + '\n'
            ret_str += get_errata_rpm_packages(url)
        ret_str += '\n'
    return ret_str


cves, nocves = get_cves()

# with open('nocve.info', 'w') as file:
#     file.write('\n'.join(nocves))
#
# with open('cve.info', 'w') as file:
#     file.write('\n'.join(cves.keys()))
keys = ['ssh', 'python', 'nginx', 'apache', 'openssl']


def check_other(cve):
    for key in keys:
        if key.lower() in str(cve).lower():
            return False
    return True


def collect_info(sub_str):
    sub_str = sub_str.lower()
    with open('cve_rpms-{}.info'.format(sub_str), 'w') as file:
        for cve, cve_num in cves.items():
            if sub_str == 'other':
                if not check_other(cve):
                    continue
            elif sub_str not in str(cve).lower():
                continue
            print(cve)
            rpm_info_str = get_errata_details(cve_num)
            file.write(cve + '：\n')
            file.write(rpm_info_str + '\n')
            file.write('\n----------------------\n')


# with open('cve_rpms.info', 'w') as file:
#     for cve, cve_num in cves.items():
#         print(cve)
#         rpm_info_str = get_errata_detail(cve_num)
#         file.write(cve + '：\n')
#         file.write(rpm_info_str + '\n')
#         file.write('\n----------------------\n')

# get_errata_details('CVE-2019-16935')

for key in keys:
    collect_info(key)

collect_info('other')
