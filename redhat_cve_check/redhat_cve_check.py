# coding=utf-8
# openpyxl require python3
# version = 1.1

import os
import json
import requests
import re
from openpyxl import load_workbook
from lxml import etree
from itertools import groupby
from typing import List

from pprint import pprint

# 抽取漏洞出现比较多的关键词,可以借助title信息进行分类
CVE_TITLE_KEYWORDS = ['ssh', 'python', 'nginx', 'apache', 'openssl']

# CVE 详情页面，和errata rpm页面，平台名称有差异，此处做个map
REDHAT_PLATFORM_NAME = {
    # <product alias name>: <affect name in cve page > ,<name for rpm packages locate>
    "rhel7": ('Red Hat Enterprise Linux 7', 'Red Hat Enterprise Linux Server 7')
}


def get_cves():
    # 打开一个workbook
    wb = load_workbook(filename="cve.xlsx")

    # 获取当前活跃的worksheet,默认就是第一个worksheet
    # ws = wb.active
    # 当然也可以使用下面的方法：获取所有表格(worksheet)的名字
    sheets = wb.sheetnames
    # 第一个表格的名称
    sheet_first = sheets[0]
    # 获取特定的worksheet
    ws = wb[sheet_first]

    # 获取表格所有行和列，两者都是可迭代的
    cves = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
        # print(line)
        cve_num = None
        title = row[1].value
        level = row[2].value
        search_obj = re.search('cve.*\d', title, re.I)
        if search_obj:
            cve_num = search_obj.group()
        title_key = next(filter(lambda x: x in title.lower(), CVE_TITLE_KEYWORDS), '未分类')
        item = {'cve_num': cve_num, 'title': title, 'level': level, 'title_key': title_key}
        cves.append(item)
    return cves


def _get_rpm_version(package_name, rpms_lists):
    for rpm in rpms_lists:
        name, version, release, _, _ = split_filename(rpm)
        if name == package_name:
            return "{}-{}-{}".format(name, version, release)


class RedHatCveCheck(object):
    headers = {
        'authority': 'access.redhat.com',
        'accept': '*/*',
        'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'dnt': '1',
        'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="102", "Google Chrome";v="102"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36'
    }

    CVE_THREAT_SEVERITY_TRANS = {
        'low': '低',
        'moderate': '中',
        'important': '高'
    }

    def __init__(self, cve_num=None, affect_product_alias='rhel7'):
        self.cve_num = cve_num
        self.platform_name_in_cve = REDHAT_PLATFORM_NAME[affect_product_alias][0]
        self.platform_name_in_errata = REDHAT_PLATFORM_NAME[affect_product_alias][1]

    def get_cve_erratas(self, get_rpm=True):
        self._check_cve_num()
        url = "https://access.redhat.com/api/redhat_node/{}.json".format(self.cve_num)
        page_url = 'https://access.redhat.com/security/cve/{}'.format(self.cve_num)
        print(page_url)
        cve_state = '红帽已处理'
        affect_products = []
        errate_package_names = []
        cve_detail = ''
        cve_statement = ''
        cve_threat_severity = ''
        response = requests.request("GET", url, headers=self.headers)
        if response.status_code == 200:
            res = response.json()
            cve_detail = res['field_cve_details_text']['und'][0]['value']
            cve_threat_severity = res['field_cve_threat_severity_text']['und'][0]['value'] \
                if res['field_cve_threat_severity_text'] else '红帽未分级'
            cve_statement = res['field_cve_statement_text']['und'][0]['value'] \
                if res['field_cve_statement_text'] else ''
            objs = res['field_cve_releases_txt']['und'][0]['object']
            objs = [obj for obj in objs if obj['product'] == self.platform_name_in_cve]
            if not objs:
                cve_state = "未找到相关修复：%s" % self.platform_name_in_cve
            for obj in objs:
                errate_url = obj['advisory'].get('url')
                products = {
                    'errate_state': obj['state'],
                    'errate_url': errate_url,
                    'errate_package_name': obj['package']
                }
                errate_package_names.append(obj['package'])
                fixed_rpms = []
                if errate_url and get_rpm:
                    fixed_rpms = self._get_errata_rpm_packages(errate_url)
                if fixed_rpms:
                    products.update(
                        {'fixed_rpms': fixed_rpms, 'rpm_version': _get_rpm_version(obj['package'], fixed_rpms)})
                affect_products.append(products)
        else:
            cve_state = '红帽未收录'
        errata_infos = {
            'cve_url': page_url,
            'cve_state': cve_state,
            'cve_threat_severity': self._trans_cve_threat_severity(cve_threat_severity),
            'cve_detail': cve_detail,
            'cve_statement': cve_statement,
            'affect_products': affect_products,
            'cve_keyword': longestCommonPrefix(errate_package_names) or '未分类'
        }
        return errata_infos

    def _get_errata_rpm_packages(self, errate_url):
        rpm_infos = None
        response = requests.request("GET", errate_url, headers=self.headers)
        html = etree.HTML(response.text)
        title_node = html.xpath('//*[@id="packages"]/h2[text()="{}"]'.format(self.platform_name_in_errata))
        if title_node and title_node[0].getnext().tag == 'table':
            table = title_node[0].getnext()
            rpm_infos = table.xpath('.//tr/td[@class="name"]//text()')
            rpm_infos = [str(s).strip() for s in rpm_infos]
        return rpm_infos

    def _check_cve_num(self):
        if not self.cve_num:
            print(type(self.cve_num), self.cve_num)
            raise ValueError("pls init cve_num")

    def set_cve(self, num):
        if not isinstance(num, str):
            print(type(num), num)
            raise ValueError('cve_num need string type')
        num = num.lower().strip()
        if not num.startswith('cve'):
            raise ValueError('cve_num need prefix "cve-"')
        self.cve_num = num

    def _trans_cve_threat_severity(self, cve_threat_severity):
        cve_threat_severity = cve_threat_severity.lower()
        return self.CVE_THREAT_SEVERITY_TRANS.get(cve_threat_severity, cve_threat_severity)


# tools
def longestCommonPrefix(strs: List[str]) -> str:
    def isCommonPrefix(length):
        str0, count = strs[0][:length], len(strs)
        return all(strs[i][:length] == str0 for i in range(1, count))

    if not strs:
        return ""

    minLength = min(len(s) for s in strs)
    low, high = 0, minLength
    while low < high:
        mid = (high - low + 1) // 2 + low
        if isCommonPrefix(mid):
            low = mid
        else:
            high = mid - 1

    return strs[0][:low]


# copy from https://github.com/rpm-software-management/yum/blob/master/rpmUtils/miscutils.py splitFilename
def split_filename(filename):
    """
    Pass in a standard style rpm fullname

    Return a name, version, release, epoch, arch, e.g.::
        foo-1.0-1.i386.rpm returns foo, 1.0, 1, i386
        1:bar-9-123a.ia64.rpm returns bar, 9, 123a, 1, ia64
    """

    if filename[-4:] == '.rpm':
        filename = filename[:-4]

    archIndex = filename.rfind('.')
    arch = filename[archIndex + 1:]

    relIndex = filename[:archIndex].rfind('-')
    rel = filename[relIndex + 1:archIndex]

    verIndex = filename[:relIndex].rfind('-')
    ver = filename[verIndex + 1:relIndex]

    epochIndex = filename.find(':')
    if epochIndex == -1:
        epoch = ''
    else:
        epoch = filename[:epochIndex]

    name = filename[epochIndex + 1:verIndex]
    return name, ver, rel, epoch, arch


def dump_to_file(filename, obj):
    with open('{}.json'.format(filename), 'w', encoding='utf8') as fp:
        fp.write(json.dumps(obj, ensure_ascii=False))


def read_from_file(filename):
    with open('{}.json'.format(filename), 'r', encoding='utf8') as fp:
        return json.load(fp=fp)


def get_cves_from_file(filename, reacquire=False):
    if not os.path.exists("{}.json".format(filename)) or reacquire:
        cves = get_cves()
        dump_to_file(filename, cves)
    return read_from_file(filename)


def get_and_dump_cve(filename='cve_infos', reacquire=False):
    # reacquire = False  # 是否重新获取
    # filename = 'cve_infos'
    cve_infos = get_cves_from_file(filename, reacquire=reacquire)
    check = RedHatCveCheck()
    # 查询并保存详情信息
    try:
        for cve in cve_infos:
            if not cve['cve_num']:
                # 该漏洞没有cve号
                continue
            elif cve.get('detail') and not reacquire:
                # 跳过已经获取过的
                continue
            check.set_cve(cve['cve_num'])
            cve['detail'] = check.get_cve_erratas(get_rpm=True)
    finally:
        dump_to_file(filename, cve_infos)
    # 分组保存
    # 1.保存无cve号的漏洞
    no_cves = list(filter(lambda x: not x.get('cve_num'), cve_infos))
    dump_to_file('no_cves', no_cves)

    # 2.分组保存cve漏洞 如果title没有分类好，需要通过cve_keyword进行分类。
    cve_infos = list(filter(lambda x: x.get('detail'), cve_infos))

    cve_infos.sort(key=lambda x: x.get('title_key'))
    groupby_cves = groupby(cve_infos, lambda x: x.get('title_key'))
    for key, group in groupby_cves:
        fname = 'cve_{}'.format(key)
        groups = list(group)
        groups.sort(key=sort_lamba, reverse=True)
        dump_to_file(fname, groups)


def sort_lamba(cve):
    cve_keyword = cve['detail']['cve_keyword']
    sort_key = re.search('rpm_version\': \'({}-.*?)\''.format(cve_keyword), str(cve))
    if sort_key:
        print(sort_key.groups(), '----')
        return sort_key.groups()[0]
    else:
        return ''


if __name__ == '__main__':
    get_and_dump_cve()
